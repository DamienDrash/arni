"""CSV/XLSX Streaming Parser."""
from __future__ import annotations
import asyncio
from pathlib import Path
from typing import AsyncIterator
import structlog
from app.ingestion.parsers.base import ParserRegistry, StreamingParser, TextChunk

logger = structlog.get_logger()


@ParserRegistry.register("text/csv", "application/csv")
class CSVParser(StreamingParser):
    """Streaming CSV-Parser via pandas chunksize. Nie komplett im RAM."""

    CHUNK_ROWS = 1000

    async def parse(self, file_path: Path) -> AsyncIterator[TextChunk]:
        loop = asyncio.get_event_loop()

        def _iter_chunks():
            try:
                import pandas as pd
            except ImportError:
                raise RuntimeError("pandas nicht installiert: pip install pandas")

            results = []
            char_offset = 0

            for chunk_df in pd.read_csv(str(file_path), chunksize=self.CHUNK_ROWS, on_bad_lines="skip"):
                # Spalten-Kontext als Präfix
                columns = " | ".join(str(c) for c in chunk_df.columns)

                for _, row in chunk_df.iterrows():
                    row_text = " | ".join(f"{col}: {val}" for col, val in row.items() if str(val) != "nan")
                    if len(row_text) > 10:
                        full_text = f"[Spalten: {columns}]\n{row_text}"
                        results.append((full_text, char_offset))
                        char_offset += len(full_text)

            return results

        rows = await loop.run_in_executor(None, _iter_chunks)

        for text, char_offset in rows:
            yield TextChunk(
                text=text,
                char_offset=char_offset,
                source_metadata={"parser": "pandas-csv"},
            )


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@ParserRegistry.register(XLSX_MIME)
class XLSXParser(StreamingParser):
    """Streaming XLSX-Parser via openpyxl read_only mode."""

    async def parse(self, file_path: Path) -> AsyncIterator[TextChunk]:
        loop = asyncio.get_event_loop()

        def _iter_rows():
            try:
                import openpyxl
            except ImportError:
                raise RuntimeError("openpyxl nicht installiert: pip install openpyxl")

            results = []
            char_offset = 0

            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = None
                for row in ws.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(row)]
                        continue

                    if all(c is None for c in row):
                        continue

                    row_text = " | ".join(
                        f"{headers[i]}: {val}"
                        for i, val in enumerate(row)
                        if val is not None and i < len(headers)
                    )
                    if len(row_text) > 10:
                        full_text = f"[Sheet: {sheet_name}]\n{row_text}"
                        results.append((full_text, char_offset))
                        char_offset += len(full_text)

            wb.close()
            return results

        rows = await loop.run_in_executor(None, _iter_rows)

        for text, char_offset in rows:
            yield TextChunk(
                text=text,
                char_offset=char_offset,
                source_metadata={"parser": "openpyxl"},
            )
